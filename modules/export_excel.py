"""
modules/export_excel.py
Экспорт данных дашборда в Excel с форматированием.
"""

from io import BytesIO
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def export_heatmap_to_excel(
    pivot_delta: pd.DataFrame,
    pivot_price: pd.DataFrame,
    pivot_model: pd.DataFrame,
    pivot_supplier: pd.DataFrame,
) -> BytesIO:
    """
    Экспортирует тепловую карту в Excel с цветовым форматированием.
    Возвращает BytesIO объект для скачивания.
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Лист 1: Тепловая карта с дельтами
        pivot_delta.to_excel(writer, sheet_name='Дельты %', index=True)
        ws_delta = writer.sheets['Дельты %']
        _format_heatmap_sheet(ws_delta, pivot_delta)
        
        # Лист 2: Цены конкурентов
        pivot_price.to_excel(writer, sheet_name='Цены конкурентов', index=True)
        ws_price = writer.sheets['Цены конкурентов']
        _format_price_sheet(ws_price)
        
        # Лист 3: Модели
        pivot_model.to_excel(writer, sheet_name='Модели', index=True)
        ws_model = writer.sheets['Модели']
        _format_text_sheet(ws_model)
        
        # Лист 4: Поставщики
        pivot_supplier.to_excel(writer, sheet_name='Поставщики', index=True)
        ws_supplier = writer.sheets['Поставщики']
        _format_text_sheet(ws_supplier)
    
    output.seek(0)
    return output


def _format_heatmap_sheet(ws, df):
    """Форматирует лист с тепловой картой дельт."""
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
    
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row, col)
            
            if cell.value is None or cell.value == "—":
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                cell.font = Font(color="999999")
                continue
            
            try:
                val = float(cell.value)
                
                if val > 15:
                    cell.fill = PatternFill(start_color="A32D2D", end_color="A32D2D", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif val > 5:
                    cell.fill = PatternFill(start_color="E24B4A", end_color="E24B4A", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                elif val > 0:
                    cell.fill = PatternFill(start_color="F09595", end_color="F09595", fill_type="solid")
                    cell.font = Font(color="791F1F")
                elif val >= -5:
                    cell.fill = PatternFill(start_color="C0DD97", end_color="C0DD97", fill_type="solid")
                    cell.font = Font(color="27500A")
                else:
                    cell.fill = PatternFill(start_color="3B6D11", end_color="3B6D11", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                
                cell.value = f"{val:+.1f}%"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            except (ValueError, TypeError):
                pass
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 15
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 12


def _format_price_sheet(ws):
    """Форматирует лист с ценами."""
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0 ₴'
                cell.alignment = Alignment(horizontal="right")
    
    ws.column_dimensions['A'].width = 15
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 12


def _format_text_sheet(ws):
    """Форматирует текстовые листы (модели, поставщики)."""
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
